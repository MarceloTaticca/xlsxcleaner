from flask import Flask, request, send_file, render_template
import pandas as pd
import io
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            return "No file uploaded", 400

        try:
            # Read the file bytes once so we can use it for both pandas and openpyxl.
            file_data = file.read()
            # Create two BytesIO streams from the same data.
            file_bytes_for_pandas = io.BytesIO(file_data)
            file_bytes_for_openpyxl = io.BytesIO(file_data)
            
            # Read only the first sheet using pandas (by default, read_excel reads the first sheet)
            df = pd.read_excel(file_bytes_for_pandas, engine='openpyxl')
        except Exception as e:
            return f"Error reading Excel file: {e}", 400

        # === Begin Data Processing (cleaning steps) ===
        # Filter rows where either 'Unnamed: 2' or 'Unnamed: 5' is not null.
        f_mask = (~df['Unnamed: 2'].isnull()) | (~df['Unnamed: 5'].isnull())
        df = df[f_mask].copy()

        # Drop columns that are completely empty (or contain only whitespace).
        cols_to_keep = [
            col for col in df.columns 
            if df[col].astype(str).str.strip().replace('nan', '').ne('').any()
        ]
        df_cleaned = df[cols_to_keep].copy()

        # Forward-fill the 'Unnamed: 5' column if it exists.
        if 'Unnamed: 5' in df_cleaned.columns:
            df_cleaned['Unnamed: 5'] = df_cleaned['Unnamed: 5'].ffill()

        # Delete header rows based on non-null values in 'Unnamed: 2'.
        f_mask = ~df_cleaned['Unnamed: 2'].isnull()
        df_cleaned = df_cleaned[f_mask].copy()

        # After filtering rows, drop again any columns that have become empty.
        cols_to_keep = [
            col for col in df_cleaned.columns 
            if df_cleaned[col].astype(str).str.strip().replace('nan', '').ne('').any()
        ]
        df_cleaned = df_cleaned[cols_to_keep]

        # Rename the columns to the expected seven names.
        # (Ensure that after filtering, there are exactly 7 columns.)
        df_cleaned.columns = [
            'data', 'plano', 'origem', 'histÛrico', 'valor', 'operaÁ„o', 'usu·rio'
        ]

        # Convert the "valor" column from Brazilian format to numeric floats.
        df_cleaned['valor'] = (
            df_cleaned['valor']
            .astype(str)
            .str.replace('.', '', regex=False)   # Remove thousand separator.
            .str.replace(',', '.', regex=False)    # Replace decimal comma with dot.
            .replace('', '0')                      # Handle empty strings.
            .astype(float)
        )

        # Convert the "data" column from text to datetime and format to YYYY-MM-DD.
        df_cleaned['data'] = pd.to_datetime(
            df_cleaned['data'], format='%d/%m/%Y', errors='coerce'
        ).dt.strftime('%Y-%m-%d')
        # === End Data Processing ===

        # Load the original workbook using openpyxl.
        try:
            wb = load_workbook(file_bytes_for_openpyxl)
        except Exception as e:
            return f"Error loading workbook with openpyxl: {e}", 400

        # (Optional) Remove an existing sheet named 'Cleaned Data' if it exists.
        if 'Cleaned Data' in wb.sheetnames:
            std = wb['Cleaned Data']
            wb.remove(std)

        # Create a new sheet for the cleaned data.
        ws = wb.create_sheet(title='Cleaned Data')
        
        # Write the DataFrame to the new sheet.
        # First, write the column headers.
        ws.append(list(df_cleaned.columns))
        # Then, write each row of data.
        for row in df_cleaned.itertuples(index=False, name=None):
            ws.append(list(row))
        
        # Save the modified workbook to an in-memory file.
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            attachment_filename="processed.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
